"""
test_export_and_audit.py -- export content correctness + audit log tests

Covers:
  POST /api/sql/export  -- TSV and XLSX content verification
  GET  /api/audit       -- audit log entries after SQL / export operations

Run from project root:
    pytest tests/test_export_and_audit.py -v
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from fastapi.testclient import TestClient

import app.main as main_module

# ---------------------------------------------------------------------------
# FIXTURE
# ---------------------------------------------------------------------------

DATASET = "aw_test_export_audit"

EXPECTED_COLUMNS = [
    "drug_name", "hcpcs_code", "total_paid", "total_claims", "service_year",
]


def _create_dataset(ds_dir: Path) -> None:
    rows = [
        {
            "drug_name":    "DrugA" if i < 50 else "DrugB",
            "hcpcs_code":   f"J{i + 1:04d}",
            "total_paid":   100.50 + i,
            "total_claims": i + 1,
            "service_year": 2023,
        }
        for i in range(100)
    ]
    # Add a row with special characters for round-trip testing
    rows.append({
        "drug_name":    'Drug "C",\twith\nnewline',
        "hcpcs_code":   "J9999",
        "total_paid":   999.99,
        "total_claims": 101,
        "service_year": 2023,
    })
    # Add a row with NULL-like values (we will query NULLs via CASE)
    df = pd.DataFrame(rows)
    df["total_paid"]   = df["total_paid"].astype("float64")
    df["total_claims"] = df["total_claims"].astype("int64")
    df["service_year"] = df["service_year"].astype("int64")
    df.to_parquet(str(ds_dir / "source.parquet"), index=False)

    meta = {
        "row_count": len(rows),
        "column_count": len(EXPECTED_COLUMNS),
        "columns": EXPECTED_COLUMNS,
        "original_type": "csv",
        "created_at": datetime.now().isoformat(),
    }
    (ds_dir / "metadata.json").write_text(json.dumps(meta), encoding="utf-8")
    (ds_dir / "_meta.json").write_text(json.dumps(meta), encoding="utf-8")


@pytest.fixture(scope="module")
def datasets_tmp(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("aw_export_audit")
    d = tmp / DATASET
    d.mkdir()
    _create_dataset(d)
    return tmp


@pytest.fixture(scope="module")
def client(datasets_tmp):
    orig_ds = main_module.DATASETS_DIR
    orig_ex = main_module.EXPORTS_DIR
    main_module.DATASETS_DIR = datasets_tmp
    main_module.EXPORTS_DIR = datasets_tmp / "_exports"
    main_module.EXPORTS_DIR.mkdir(exist_ok=True)
    # Clear any pre-existing audit log so tests start clean
    audit_path = datasets_tmp / "_audit.jsonl"
    if audit_path.exists():
        audit_path.unlink()
    with TestClient(main_module.app) as c:
        yield c
    main_module.DATASETS_DIR = orig_ds
    main_module.EXPORTS_DIR = orig_ex


# ===========================================================================
# TSV EXPORT — content correctness
# ===========================================================================

# Prevents TSV column headers from diverging from SELECT columns
def test_tsv_export_column_headers_match_select(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name, total_paid, total_claims FROM dataset LIMIT 5",
        "format": "tsv",
    })
    assert resp.status_code == 200
    lines = resp.content.decode("utf-8").splitlines()
    headers = lines[0].split("\t")
    assert headers == ["drug_name", "total_paid", "total_claims"]


# Prevents TSV export from returning wrong number of data rows
def test_tsv_export_row_count_matches_query(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT * FROM dataset LIMIT 10",
        "format": "tsv",
    })
    assert resp.status_code == 200
    lines = [l for l in resp.content.decode("utf-8").splitlines() if l.strip()]
    # header + 10 data rows
    assert len(lines) == 11


# Prevents data values from being garbled in TSV export
def test_tsv_export_first_row_data_values_correct(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name, total_paid, total_claims FROM dataset ORDER BY total_claims LIMIT 1",
        "format": "tsv",
    })
    assert resp.status_code == 200
    lines = resp.content.decode("utf-8").splitlines()
    fields = lines[1].split("\t")
    assert fields[0] == "DrugA"
    assert float(fields[1]) == 100.50
    assert int(fields[2]) == 1


# Prevents TSV from silently using a wrong delimiter
def test_tsv_export_uses_tab_delimiter(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name, total_paid FROM dataset LIMIT 3",
        "format": "tsv",
    })
    assert resp.status_code == 200
    header_line = resp.content.decode("utf-8").splitlines()[0]
    # Must contain tabs
    assert "\t" in header_line
    # Must not be comma-delimited or pipe-delimited for the header
    # (the header has exactly 2 columns so exactly 1 tab)
    assert header_line.count("\t") == 1


# Prevents WHERE filter from being ignored in export
def test_tsv_export_with_where_filter_reduces_rows(client):
    # Unfiltered: 101 total rows; DrugA: 50 rows
    resp_filtered = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT * FROM dataset WHERE drug_name = 'DrugA'",
        "format": "tsv",
    })
    assert resp_filtered.status_code == 200
    filtered_lines = [l for l in resp_filtered.content.decode("utf-8").splitlines() if l.strip()]

    resp_all = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT * FROM dataset",
        "format": "tsv",
    })
    assert resp_all.status_code == 200
    all_lines = [l for l in resp_all.content.decode("utf-8").splitlines() if l.strip()]

    # Filtered must have fewer rows than unfiltered
    assert len(filtered_lines) < len(all_lines)
    # DrugA has exactly 50 rows -> 51 lines (header + 50)
    assert len(filtered_lines) == 51


# ===========================================================================
# XLSX EXPORT — content correctness
# ===========================================================================

# Prevents XLSX column headers from diverging from SELECT columns
def test_xlsx_export_column_headers_match_select(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name, total_paid, total_claims FROM dataset LIMIT 5",
        "format": "xlsx",
    })
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == ["drug_name", "total_paid", "total_claims"]


# Prevents XLSX export from returning wrong number of data rows
def test_xlsx_export_row_count_matches_query(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT * FROM dataset LIMIT 10",
        "format": "xlsx",
    })
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # ws.max_row includes header row
    assert ws.max_row == 11  # 1 header + 10 data


# Prevents data values from being garbled in XLSX export
def test_xlsx_export_spot_check_data_value(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name, total_paid FROM dataset ORDER BY total_claims LIMIT 1",
        "format": "xlsx",
    })
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Row 2 is first data row (row 1 is header)
    assert ws.cell(row=2, column=1).value == "DrugA"
    assert ws.cell(row=2, column=2).value == 100.50


# ===========================================================================
# EXPORT — special characters and NULLs
# ===========================================================================

# Prevents special characters from being corrupted in export round-trip
def test_tsv_export_special_characters_survive(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name FROM dataset WHERE hcpcs_code = 'J9999'",
        "format": "tsv",
    })
    assert resp.status_code == 200
    # pandas to_csv with sep=\t will quote fields containing special chars
    content = resp.content.decode("utf-8")
    lines = content.splitlines()
    assert len(lines) >= 2
    # The value should be recoverable -- read it back with pandas
    df = pd.read_csv(io.StringIO(content), sep="\t")
    val = df.iloc[0]["drug_name"]
    assert '"C"' in val, f"Double quotes not preserved: {val!r}"
    assert "," in val, f"Comma not preserved: {val!r}"


# Prevents NULL values from being exported as the string "None"
def test_tsv_export_null_values_not_none_string(client):
    resp = client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": (
            "SELECT drug_name, "
            "CASE WHEN drug_name = 'DrugB' THEN total_paid ELSE NULL END AS paid "
            "FROM dataset LIMIT 10"
        ),
        "format": "tsv",
    })
    assert resp.status_code == 200
    content = resp.content.decode("utf-8")
    # pandas to_csv writes NaN/None as empty string by default
    # It should NOT contain the literal string "None"
    df = pd.read_csv(io.StringIO(content), sep="\t")
    # DrugA rows should have NaN (not the string "None")
    drug_a_rows = df[df["drug_name"] == "DrugA"]
    assert len(drug_a_rows) > 0
    for val in drug_a_rows["paid"]:
        assert pd.isna(val), f"Expected NaN for NULL, got {val!r}"


# ===========================================================================
# AUDIT LOG
# ===========================================================================

# Prevents audit log from being empty after SQL execution
def test_audit_has_entries_after_sql(client):
    # Execute a SQL query to generate an audit entry
    client.post("/api/sql", json={
        "dataset": DATASET,
        "sql": "SELECT COUNT(*) AS n FROM dataset",
    })
    resp = client.get("/api/audit")
    assert resp.status_code == 200
    events = resp.json()["events"]
    assert len(events) >= 1


# Prevents audit entries from missing required fields
def test_audit_entry_has_required_fields(client):
    # Execute a SQL query
    client.post("/api/sql", json={
        "dataset": DATASET,
        "sql": "SELECT COUNT(*) AS n FROM dataset",
    })
    resp = client.get("/api/audit")
    events = resp.json()["events"]
    # Find a sql-type event
    sql_events = [e for e in events if e.get("event") == "sql"]
    assert len(sql_events) >= 1, f"No sql events found in: {events}"
    entry = sql_events[0]
    assert "ts" in entry, "Audit entry missing 'ts' (timestamp)"
    assert "event" in entry, "Audit entry missing 'event' type"
    assert "dataset" in entry, "Audit entry missing 'dataset' name"
    assert entry["dataset"] == DATASET


# Prevents audit log from losing entries after multiple executions
def test_audit_multiple_entries_after_multiple_queries(client):
    # Get baseline count
    baseline = len(client.get("/api/audit").json()["events"])

    # Execute two more queries
    client.post("/api/sql", json={
        "dataset": DATASET,
        "sql": "SELECT drug_name FROM dataset LIMIT 1",
    })
    client.post("/api/sql", json={
        "dataset": DATASET,
        "sql": "SELECT total_paid FROM dataset LIMIT 1",
    })

    events = client.get("/api/audit").json()["events"]
    assert len(events) >= baseline + 2


# Prevents audit log from being returned in wrong order
def test_audit_preserves_chronological_order(client):
    # The /api/audit endpoint returns newest first
    events = client.get("/api/audit").json()["events"]
    sql_events = [e for e in events if e.get("event") == "sql" and "ts" in e]
    if len(sql_events) >= 2:
        # Newest first means timestamps should be descending
        for i in range(len(sql_events) - 1):
            assert sql_events[i]["ts"] >= sql_events[i + 1]["ts"], (
                f"Audit events not in descending chronological order: "
                f"{sql_events[i]['ts']} < {sql_events[i + 1]['ts']}"
            )


# Prevents export events from being missing in audit log
def test_audit_records_export_event(client):
    # Perform an export
    client.post("/api/sql/export", json={
        "dataset": DATASET,
        "sql": "SELECT * FROM dataset LIMIT 5",
        "format": "tsv",
    })
    events = client.get("/api/audit").json()["events"]
    export_events = [e for e in events if e.get("event") == "sql_export"]
    assert len(export_events) >= 1, f"No sql_export events found in audit log"
    assert export_events[0]["status"] == "success"
    assert export_events[0]["dataset"] == DATASET
